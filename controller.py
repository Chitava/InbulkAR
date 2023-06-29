import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
import view
from tkinter import filedialog as fd
import db
import os
import xlwt
import xlsxwriter
import math
import datetime
from datetime import time



def Workers():
    workers = []
    temp = db.Read_workers()
    for item in temp:
        workers.append(item[0])
    return workers


def Set_date():
    return 0

def Create_db():
    file = fd.askopenfilename()
    wb = load_workbook(file)
    sheet = wb.worksheets[0]
    ar = {}  # общий массив данных
    result = []
    temp = []
    for i in range(1, sheet.max_row + 1):
        for j in range(2, 20):
            data = sheet.cell(row=i, column=j).value
            if (data == '' or data is None or data == '\n' or 'parsec' in data or '/' in data):
                continue
            if data == '--\n--\n--':
                temp.append('0,0,0')
            else:
                temp.append(data.replace("\n", ', ').replace(':', '.').replace(' ', ''))
        if i%2 != 0:
            continue
        else:
            if len(temp)>0:
                result.append(list(temp))
        temp.clear()
    for item in result:
        item[0] = item[0].replace(',', ' ')
        for i in range(1, len(item)):
            item[i] = (item[i].split(',')).pop(2).replace('--', '0')

        key = item[0]
        item.pop(0)
        for i in range(len(item)):
            if len(item[i]) == 0:
                item.pop(i)
                item.insert(i, 0)
        if (len(item) < 31):
            for i in range(len(item), 31):
                item.append("0")
        ar[key] = item
        #sort = dict(sorted(ar.items()))
    return ar


def Insert_db_data():
    ar = Create_db()
    add_workers = []
    db.Create_table_workers()
    for key, val in ar.items():
        worker = db.Find_worker(key)
        if len(worker) == 0:
            add_workers.append(key)
    if len(add_workers) > 0:
        view.Add_workers_form(add_workers)
    if (db.Find_db(view.db_date) == -1):
        db.Create_new_table_work_day(view.db_date)
    for key, val in ar.items():
        print(db.Select_work_days_for_one_worker(view.db_date, key))
        if (db.Select_work_days_for_one_worker(view.db_date, key) == None or len(db.Select_work_days_for_one_worker
                                                                                     (view.db_date, key)) == 0):
            db.Input_new_workdays(view.db_date, key, val)
        else:
            db.Update_new_workdays(view.db_date, key, val)
    view.Messagebox("Запись в базу данных", f"Новые данные за {view.db_date} внесены успешно")


def Create_salary_in_one_month(date1, date2, month):
    date1 = int(date1)
    date2 = int(date2)
    salary_all = {}
    result = []
    data = db.Select_work_days_join_workers(month)
    ld=len(data)
    for item in data:
        name = str(item[0])
        work_time = datetime.datetime(2000, 1, 1, 9, 0)
        work_day = 0
        salary = 0
        elabor_time = datetime.datetime(2000, 1, 1, 0, 0)
        elabor_salary = 0
        all_elab_time = datetime.datetime(2000, 1, 1, 0, 0)
        wage = item[1]
        elab = item[2]
        houre_wage = wage / 8
        result = []
        for i in range(3+date1, 4+date2):
            temp = item[i].split('.')
            if len(temp) > 1:
                curent_time = datetime.datetime(2000, 1, 1, int(temp[0]), int(temp[1]))
                if curent_time.hour < work_time.hour:
                    day_wage = (float(item[i]) - 1) * houre_wage
                    salary += day_wage
                    work_day += 1
                else:
                    salary += wage
                    elab_time = curent_time - work_time
                    all_elab_time += elab_time
                    work_day += 1
        temp_elab_time = (all_elab_time.day - 1) * 24
        month_salary = salary + (float(f"{temp_elab_time + all_elab_time.hour}.{all_elab_time.minute}") * elab)
        result.append(work_day)  # Кол-во отработаных дней
        result.append(f"{temp_elab_time + all_elab_time.hour},{all_elab_time.minute}")  # Сумма часов переработки
        result.append(round(salary, 2))  # Зарплата за отработанные дни
        result.append(
            float(f"{temp_elab_time + all_elab_time.hour}.{all_elab_time.minute}") * elab)  # Зарплата за переработку
        result.append(round(month_salary, 2))  # Зарплата за месяц
        salary_all[name] = result
    return salary_all


def Create_salary_with_last_month(date1, date2, last_month, month):
    date1 = int(date1)
    date2 = int(date2)
    salary = {}
    now_work_days = db.Select_work_days(month)
    last_work_days = db.Select_work_days(last_month)
    all_workers = db.Read_workers()
    for name in all_workers:
        work_day = 0
        sal = 0
        elab_day = 0
        elab_time = 0
        sal_elabor = 0
        wage = 0
        elabor = 0
        res = []
        for day in last_work_days:
            temporery = []
            if name[0] == day[0]:
                work_day = 0
                sal = 0
                elab_day = 0
                elab_time = 0
                sal_elabor = 0
                wage = 0
                elabor = 0
                res = []
                for i in range(int(date1), len(day)):
                    if float(day[i]) ==0:
                        wage = 0
                    elif float(day[i]) > 9:
                        elabor = (float(day[i]) - 9) * float(name[2])
                        wage = int(name[1])
                        work_day += 1
                        elab_day += 1
                        elab_time += (float(day[i]) - 9)
                        sal_elabor+= (float(day[i]) - 9)*name[2]
                    elif float(day[i]) <= 9:
                        wage = (int(name[1])/8)*(float(day[i])-1)
                        work_day+=1
                    sal += round(wage, 2)
        for day in now_work_days:
            if name[0] == day[0]:
                for i in range(1, int(date2)+1):
                    if float(day[i]) == 0:
                        wage = 0
                    elif float(day[i]) > 9:
                        elabor = (float(day[i]) - 9) * float(name[2])
                        wage = name[1]
                        work_day += 1
                        elab_day += 1
                        elab_time += (float(day[i]) - 9)
                        sal_elabor += (float(day[i]) - 9) * name[2]
                    elif float(day[i]) <= 9:
                        wage = (int(name[1])/8) * (float(day[i]) - 1)
                        work_day += 1
                    sal += round(wage, 2)
        res = []
        res.append(work_day)
        res.append(round(sal, 2))
        res.append(elab_day)
        res.append(round(elab_time, 2))
        res.append(round(sal_elabor, 2))
        salary[name[0]] = res
        Write_salary_worker(name[0], round(sal+sal_elabor, 2), month)
    return salary

def Create_salary_for_one(data):
    date1 = int(date1)
    date2 = int(date2)
    salary = {}
    wag = db.Read_worker(data[0])
    temp_days = data[1].split(';')
    work_day = 0
    sal = 0
    elab_day = 0
    elab_time = 0
    sal_elabor = 0
    for i in range(len(temp_days)):
        wage = 0
        elabor = 0
        res =[]
        if (temp_days[i] != '--,--,--' and len(temp_days[i])>2):
            times = temp_days[i].split(",")
            if times[2] != '--':
                work_day += 1
                if round(float(times[2])) <= 9:
                    wage = ((int(wag[0][1]))/8) * float(times[2])
                else:
                    elab_day+=1
                    wage = int(wag[0][1])
                    elabor = (round(float(times[2])) - 9)*(int(wag[0][2]))
                    elab_time += (float(times[2]) - 9)
                    sal_elabor += elabor
        sal += wage
        res = []
        res.append(work_day)
        res.append(round(sal, 2))
        res.append(elab_day)
        res.append(round(elab_time, 2))
        res.append(round(sal_elabor, 2))
        salary[data[0]] = res
    return salary


#

def Write_salary_worker(name, salary, month):
    salary_workers = []
    salary_workers = db.Read_salary_workers(month)
    if salary_workers == 0 or len(salary_workers)==0:
        db.Create_table_salarys(month)
        db.Add_salary_worker(name, salary, month)
    else:
        count = 0
        for item in salary_workers:
            if name == item[0]:
                db.Update_salary_workers(name, salary, month)
                count+=1
                break
        if count == 0:
           db.Add_salary_worker(name, salary, month)



def Read_last_month_salary(name):
    last_month = []
    date_now = view.db_date.split()

    if date_now[0] == '01':
        last_month.append(12)
        last_month.append(int(date_now[1])-1)
    else:
        str_date = '0'+str(int(date_now[0])-1)
        last_month.append(str_date)
        last_month.append(date_now[1])
        last_month = " ".join(last_month)
        last_month_salarys = db.Read_salary_workers(last_month)
        if last_month_salarys == 1 or last_month_salarys == 0:
            return 0
        else:
            for item in last_month_salarys:
                if item[0] == name:
                    if item[1] < 0:
                        return item[1]*(-1)
                    else:
                        return item[1]


def Print_all(data):
    for keys, val in data.items():
        file.write(f"{keys}\n\n")
        last_month_salary = 0
        for item in last:
            if keys == item[0]:
                if item[1] < 0:
                    avan=item[1]*(-1)
        file.write(f"Долг за прошлый месяц: {0}\n")
        file.write(f"Отработано: - {val[2]} дней\n")
        file.write(f"Заработал: {val[1]} р.\n")
        file.write(f"Аванс: {0} р.\n ")
        file.write(f"Часов переработки: {val[3]}\n")
        file.write(f"Заработал за перерработку: {val[4]} р.\n")
        file.write(f"Итого за месяц - {round(float(val[2])+float(val[4]), 2)} р.\n\n")
        file.write("******************************************************************\n")
    file.close()
    os.startfile("temp.txt", "print")



def Save_to_excel(salarys, month):

    book = xlsxwriter.Workbook(f"Зарплата за {month}.xlsx")
    sheet = book.add_worksheet("Зарплата")
    row = 0
    column = 0
    content = ["№", "Имя", "Рабочие дни", "Часы перерработки", "Зарплата за месяц", "Зарплата за переработку", "Аванс",
               "Зарплата за месяц", "Итого с учетом аванса"]
    for item in content:
        sheet.write(row, column, item)
        column += 1
    row = 1
    column = 0
    result = []
    for keys, val in salarys.items():
        values = []
        values.append(row)
        values.append(keys)
        values.append(val[0])
        values.append(val[1])
        values.append(val[2])
        values.append(val[3])
        values.append(0)
        values.append(val[4])
        values.append((f'=H{row + 1}-G{row + 1}'))
        result.append(values)
        row+=1
    row = 1
    for item in result:
        column = 0
        for val in item:
            sheet.write(row, column, val)
            column += 1
        row+=1
    book.close()


def Create_salary_in_one_month_for_one(name, date1, date2, month):
    salary = []
    data = db.Select_work_days_for_one_worker(name, month)
    date1 = int(date1)
    date2 = int(date2)
    salary_all = {}
    result = []
    for item in data:
        name = str(item[0])
        work_time = datetime.datetime(2000, 1, 1, 9, 0)
        work_day = 0
        salary = 0
        elabor_time = datetime.datetime(2000, 1, 1, 0, 0)
        elabor_salary = 0
        all_elab_time = datetime.datetime(2000, 1, 1, 0, 0)
        wage = item[1]
        elab = item[2]
        houre_wage = wage / 8
        result = []
        for i in range(3 + date1, 4 + date2):
            temp = item[i].split('.')
            if len(temp) > 1:
                curent_time = datetime.datetime(2000, 1, 1, int(temp[0]), int(temp[1]))
                if curent_time.hour < work_time.hour:
                    day_wage = (float(item[i]) - 1) * houre_wage
                    salary += day_wage
                    work_day += 1
                else:
                    salary += wage
                    elab_time = curent_time - work_time
                    all_elab_time += elab_time
                    work_day += 1
        temp_elab_time = (all_elab_time.day - 1) * 24
        month_salary = salary + (float(f"{temp_elab_time + all_elab_time.hour}.{all_elab_time.minute}") * elab)
        result.append(work_day)  # Кол-во отработаных дней
        result.append(f"{temp_elab_time + all_elab_time.hour},{all_elab_time.minute}")  # Сумма часов переработки
        result.append(round(salary, 2))  # Зарплата за отработанные дни
        result.append(
            float(f"{temp_elab_time + all_elab_time.hour}.{all_elab_time.minute}") * elab)  # Зарплата за переработку
        result.append(round(month_salary, 2))  # Зарплата за месяц
        salary_all[name] = result
    return salary_all


def Create_salary_in_two_month_for_one(name, day1, day2, first_month, second_month):
    sum_sallary = {}
    sum_sallary_val =[]
    sallary_first_month = Create_salary_in_one_month_for_one(name, day1, '31', first_month)
    sallary_second_month = Create_salary_in_one_month_for_one(name, '01', day2, second_month)
    first_month_val = sallary_first_month[name]
    second_month_val = sallary_second_month[name]
    for i in range(len(first_month_val)):
        sum_sallary_val.append(first_month_val[i] + second_month_val[i])
    sum_sallary[name] = sum_sallary_val
    return sum_sallary

def Create_salary_in_two_month_for_all(day1, day2, first_month, second_month):
    sum_sallary_now ={}
    sallary_first_month = Create_salary_in_one_month(day1, "31", first_month)
    sallary_second_month = Create_salary_in_one_month('01', day2, second_month)
    for key_first, val_first in sallary_first_month.items():
        for key_last, val_last in sallary_second_month.items():
            if key_first == key_last:
                sum_sallary = []
                for i in range(len(val_first)):
                    sum_sallary.append(val_first[i] + val_last[i])
                    sum_sallary_now[key_first] = sum_sallary
                break
    return sum_sallary_now


