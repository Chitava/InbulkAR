import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
import view
from tkinter import filedialog as fd
import db
import os
import xlwt
import xlsxwriter


def Create_db():
    file = fd.askopenfilename()
    wb = load_workbook(file)
    sheet = wb.worksheets[0]
    ar = {}  # общий массив данных
    result = []
    temp = []
    for i in range(6, sheet.max_row + 1):
        for j in range(1, 21):
            data = sheet.cell(row=i, column=j).value
            if (data == '' or data is None or data == '\n'):
                continue
            if data == '--\n--\n--':
                temp.append('--,--,--')
            else:
                temp.append(data.replace("\n", ', ').replace(':', '.').replace(' ', ''))
        if len(temp)>0:
            result.append(list(temp))
        temp.clear()
    for item in result:
        for val in item:
            if 'parsec' in val:
                result.remove(item)
    for i in range(len(result) - 1):
        if i % 2 == 0:
            temp.append(result[i] + result[i + 1])
    for i in range(len(temp)):
        if len(temp[i]) == 0:
            continue
        else:
            temp_data = []
            keys = temp[i][1].replace(',', ' ')
            temp[i].pop(0)
            temp[i].pop(0)
            for item in temp[i]:
                temp_data.append(item)
            if (len(temp_data) < 31):
                lenth_data = 31 - len(temp_data)
                for i in range(lenth_data):
                    temp_data.append('--,--,--')
            ar[keys] = temp_data
    return ar

ar = Create_db()
for key, val in ar.items():
    print(key)
    print(val)
